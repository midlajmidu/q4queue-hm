import os
import uuid
import pandas as pd
from datetime import datetime, timezone, timedelta
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func
from app.models.export_job import ExportJob
from app.models.user import User
from app.models.organization import Organization
from app.models.queue import Queue
from app.models.session import Session
from app.models.token import Token
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from app.models.parent_organization import ParentOrganization
async def _resolve_date_bounds(db: AsyncSession, job: ExportJob):
    parent_org = await db.get(ParentOrganization, job.parent_org_id)
    tz_name = parent_org.timezone if parent_org and parent_org.timezone else "UTC"
    
    try:
        from zoneinfo import ZoneInfo
        tz = ZoneInfo(tz_name)
    except Exception:
        tz = timezone.utc

    now = datetime.now(tz)
    start_date = None
    end_date = None

    date_range = job.filters.get("date_range", "All Time")
    
    if date_range == "Today":
        start_date = now.date()
        end_date = now.date()
    elif date_range == "Yesterday":
        start_date = (now - timedelta(days=1)).date()
        end_date = start_date
    elif date_range == "Last 7 Days":
        start_date = (now - timedelta(days=6)).date()
        end_date = now.date()
    elif date_range == "Last 30 Days":
        start_date = (now - timedelta(days=29)).date()
        end_date = now.date()
    elif date_range == "This Month":
        start_date = now.replace(day=1).date()
        end_date = now.date()
    elif date_range == "Last Month":
        last_day = now.replace(day=1) - timedelta(days=1)
        start_date = last_day.replace(day=1).date()
        end_date = last_day.date()
    elif date_range == "Custom Date Range":
        c_start = job.filters.get("custom_start_date")
        c_end = job.filters.get("custom_end_date")
        if c_start:
            start_date = datetime.strptime(c_start.split("T")[0], "%Y-%m-%d").date()
        if c_end:
            end_date = datetime.strptime(c_end.split("T")[0], "%Y-%m-%d").date()

    return start_date, end_date, tz

async def _get_org_ids(db: AsyncSession, parent_org_id: uuid.UUID, filters: dict):
    stmt = select(Organization).where(Organization.parent_organization_id == parent_org_id)
    branch_ids = filters.get("branch_ids")
    if branch_ids:
        stmt = stmt.where(Organization.id.in_([uuid.UUID(b) for b in branch_ids]))
    
    res = await db.execute(stmt)
    orgs = res.scalars().all()
    return [o.id for o in orgs], {o.id: o.name for o in orgs}, [o for o in orgs]

def _apply_date_filter(stmt, start_date, end_date, tz):
    if start_date:
        start_dt = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=tz)
        stmt = stmt.where(Token.created_at >= start_dt)
    if end_date:
        end_dt = datetime.combine(end_date, datetime.max.time()).replace(tzinfo=tz)
        stmt = stmt.where(Token.created_at <= end_dt)
    return stmt


async def _generate_structured_excel(sections: list, file_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    section_font = Font(size=14, bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    row_idx = 1
    for section in sections:
        title = section.get('title')
        data = section.get('data') # List of dicts
        
        if title:
            cell = ws.cell(row=row_idx, column=1, value=title)
            cell.font = section_font
            row_idx += 2
            
        if not data:
            ws.cell(row=row_idx, column=1, value="No data available")
            row_idx += 2
            continue
            
        headers = list(data[0].keys())
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        row_idx += 1
        
        for row_data in data:
            for col_idx, h in enumerate(headers, 1):
                val = row_data.get(h)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
            row_idx += 1
            
        row_idx += 2 # Space between sections
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)
        
    wb.save(file_path)


async def _generate_csv(sections: list, file_path: str):
    import csv
    with open(file_path, mode='w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for section in sections:
            title = section.get('title')
            data = section.get('data')
            if title:
                writer.writerow([f"=== {title.upper()} ==="])
            if not data:
                writer.writerow(["No data available"])
                writer.writerow([])
                continue
                
            headers = list(data[0].keys())
            writer.writerow(headers)
            for row in data:
                writer.writerow([row.get(h, "") for h in headers])
            writer.writerow([])


def _format_time(seconds: float) -> str:
    if not seconds or seconds < 0:
        return "0m"
    mins = int(seconds // 60)
    if mins >= 60:
        hrs = mins // 60
        rem_mins = mins % 60
        return f"{hrs}h {rem_mins}m"
    return f"{mins}m"


async def _report_executive_summary(job: ExportJob, db: AsyncSession, file_path: str, format_type: str):
    start_date, end_date, tz = await _resolve_date_bounds(db, job)
    org_ids, org_name_map, orgs = await _get_org_ids(db, job.parent_org_id, job.filters)
    
    if not org_ids:
        return [{"Message": "No branches found."}]
        
    from sqlalchemy.dialects.postgresql import aggregate_order_by
    from app.models.token import TokenStatus
    
    stmt = select(
        Token.org_id,
        func.count(Token.id).filter(Token.status != TokenStatus.deleted).label("total"),
        func.count(Token.id).filter(Token.status == TokenStatus.done).label("served"),
        func.count(Token.id).filter(Token.status == TokenStatus.waiting).label("waiting"),
        func.count(Token.id).filter(Token.status == TokenStatus.skipped).label("cancelled"),
        func.count(Token.id).filter(Token.status == TokenStatus.deleted).label("removed"),
        func.avg(func.extract('epoch', Token.served_at - Token.created_at)).filter(Token.status == TokenStatus.done).label("avg_wait"),
        func.max(func.extract('epoch', Token.served_at - Token.created_at)).filter(Token.status == TokenStatus.done).label("max_wait"),
        func.avg(func.extract('epoch', Token.completed_at - Token.served_at)).filter(Token.status == TokenStatus.done).label("avg_serve"),
        func.max(func.extract('epoch', Token.completed_at - Token.served_at)).filter(Token.status == TokenStatus.done).label("max_serve")
    ).where(Token.org_id.in_(org_ids))
    
    stmt = _apply_date_filter(stmt, start_date, end_date, tz)
    stmt = stmt.group_by(Token.org_id)
    
    res = await db.execute(stmt)
    branch_metrics = res.fetchall()
    
    # Query 2: Staff counts
    staff_stmt = select(
        User.org_id,
        func.count(User.id).label("total"),
        func.sum(func.cast(User.is_active, func.integer())).label("active")
    ).where(User.org_id.in_(org_ids), User.role.in_(["admin", "staff"])).group_by(User.org_id)
    staff_res = await db.execute(staff_stmt)
    staff_counts = {row.org_id: {"total": row.total, "active": row.active} for row in staff_res.fetchall()}
    
    # Compute Aggregates
    tot_branches = len(orgs)
    act_branches = sum(1 for o in orgs if o.is_active)
    tot_staff = sum(s["total"] or 0 for s in staff_counts.values())
    act_staff = sum(s["active"] or 0 for s in staff_counts.values())
    
    tot_cust = sum(r.total for r in branch_metrics)
    tot_served = sum(r.served for r in branch_metrics)
    tot_waiting = sum(r.waiting for r in branch_metrics)
    tot_cancelled = sum(r.cancelled for r in branch_metrics)
    tot_removed = sum(r.removed for r in branch_metrics)
    
    avg_wait = sum((r.avg_wait or 0) * r.served for r in branch_metrics) / tot_served if tot_served else 0
    max_wait = max((r.max_wait for r in branch_metrics if r.max_wait), default=0)
    avg_serve = sum((r.avg_serve or 0) * r.served for r in branch_metrics) / tot_served if tot_served else 0
    max_serve = max((r.max_serve for r in branch_metrics if r.max_serve), default=0)
    
    serve_rate = round((tot_served / tot_cust) * 100, 1) if tot_cust else 0
    
    overview_data = [
        {"Metric": "Total Branches", "Value": tot_branches},
        {"Metric": "Active Branches", "Value": act_branches},
        {"Metric": "Total Staff", "Value": tot_staff},
        {"Metric": "Active Staff", "Value": act_staff},
        {"Metric": "Total Customers", "Value": tot_cust},
        {"Metric": "Customers Served", "Value": tot_served},
        {"Metric": "Customers Waiting", "Value": tot_waiting},
        {"Metric": "Cancelled", "Value": tot_cancelled},
        {"Metric": "Service Completion Rate", "Value": f"{serve_rate}%"},
        {"Metric": "Average Waiting Time", "Value": _format_time(avg_wait)},
        {"Metric": "Maximum Waiting Time", "Value": _format_time(max_wait)},
        {"Metric": "Average Service Time", "Value": _format_time(avg_serve)},
        {"Metric": "Maximum Service Time", "Value": _format_time(max_serve)},
    ]
    
    # Top 5 branches by volume
    branch_detail = []
    for r in branch_metrics:
        b_name = org_name_map.get(r.org_id, "Unknown")
        rate = round((r.served / r.total) * 100, 1) if r.total else 0
        branch_detail.append({
            "Branch": b_name,
            "Total Customers": r.total,
            "Customers Served": r.served,
            "Service Completion %": f"{rate}%",
            "Avg Waiting Time": _format_time(r.avg_wait or 0),
            "Avg Service Time": _format_time(r.avg_serve or 0),
            "score": (rate * 0.5) + ((100 - min((r.avg_wait or 0)/60, 100)) * 0.3) + (min(r.total/100, 1) * 20)
        })
    
    branch_detail.sort(key=lambda x: x["Total Customers"], reverse=True)
    
    # Staff performance
    top_staff_stmt = select(
        User.first_name, User.last_name, User.email,
        func.count(Token.id).label("served"),
        func.avg(func.extract('epoch', Token.completed_at - Token.served_at)).label("avg_serve")
    ).join(Token, Token.completed_by_id == User.id).where(
        User.org_id.in_(org_ids), Token.status == TokenStatus.done
    )
    top_staff_stmt = _apply_date_filter(top_staff_stmt, start_date, end_date, tz)
    top_staff_stmt = top_staff_stmt.group_by(User.id).order_by(func.count(Token.id).desc()).limit(5)
    
    staff_res = await db.execute(top_staff_stmt)
    top_staff = []
    for row in staff_res.fetchall():
        name = f"{row.first_name or ''} {row.last_name or ''}".strip() or row.email.split('@')[0]
        top_staff.append({
            "Staff Name": name,
            "Customers Served": row.served,
            "Avg Service Time": _format_time(row.avg_serve or 0)
        })
        
    # Key Insights
    best_branch = max(branch_detail, key=lambda x: x["score"]) if branch_detail else None
    highest_vol = branch_detail[0] if branch_detail else None
    lowest_wait = min((b for b in branch_detail if b["Total Customers"] > 0), key=lambda x: x.get("_raw_wait", float(x["Avg Waiting Time"].split('m')[0].replace('h ', ''))), default=None) if branch_detail else None
    needs_attn = min(branch_detail, key=lambda x: x["score"]) if branch_detail else None
    
    insights = []
    if best_branch: insights.append({"Insight": "Highest Performing Branch", "Value": best_branch["Branch"]})
    if highest_vol: insights.append({"Insight": "Highest Customer Volume", "Value": highest_vol["Branch"]})
    if lowest_wait: insights.append({"Insight": "Lowest Average Waiting Time", "Value": lowest_wait["Branch"]})
    insights.append({"Insight": "Average Customers per Branch", "Value": round(tot_cust / act_branches) if act_branches else 0})
    insights.append({"Insight": "Average Customers per Staff", "Value": round(tot_cust / act_staff) if act_staff else 0})
    insights.append({"Insight": "Service Completion Rate", "Value": f"{serve_rate}%"})
    if needs_attn: insights.append({"Insight": "Branch Requiring Attention", "Value": needs_attn["Branch"]})
    
    # Remove score from branch detail for output
    for b in branch_detail:
        b.pop("score", None)
        
    sections = [
        {"title": "Overview KPIs", "data": overview_data},
        {"title": "Key Insights", "data": insights},
        {"title": "Top 5 Busiest Branches", "data": branch_detail[:5]},
        {"title": "Top 5 Best Performing Staff", "data": top_staff},
        {"title": "Branch Comparison", "data": branch_detail}
    ]
    
    if format_type == "CSV":
        await _generate_csv(sections, file_path)
    else:
        await _generate_structured_excel(sections, file_path)

async def _report_branch_performance(job: ExportJob, db: AsyncSession, file_path: str, format_type: str):
    start_date, end_date, tz = await _resolve_date_bounds(db, job)
    org_ids, org_name_map, orgs = await _get_org_ids(db, job.parent_org_id, job.filters)
    
    if not org_ids:
        return [{"Message": "No branches found."}]
        
    from app.models.token import TokenStatus
    
    # Get Branch Metrics
    stmt = select(
        Token.org_id,
        func.count(Token.id).filter(Token.status != TokenStatus.deleted).label("total"),
        func.count(Token.id).filter(Token.status == TokenStatus.done).label("served"),
        func.count(Token.id).filter(Token.status == TokenStatus.waiting).label("waiting"),
        func.count(Token.id).filter(Token.status == TokenStatus.skipped).label("cancelled"),
        func.avg(func.extract('epoch', Token.served_at - Token.created_at)).filter(Token.status == TokenStatus.done).label("avg_wait"),
        func.max(func.extract('epoch', Token.served_at - Token.created_at)).filter(Token.status == TokenStatus.done).label("max_wait"),
        func.avg(func.extract('epoch', Token.completed_at - Token.served_at)).filter(Token.status == TokenStatus.done).label("avg_serve"),
        func.max(func.extract('epoch', Token.completed_at - Token.served_at)).filter(Token.status == TokenStatus.done).label("max_serve")
    ).where(Token.org_id.in_(org_ids))
    stmt = _apply_date_filter(stmt, start_date, end_date, tz)
    stmt = stmt.group_by(Token.org_id)
    branch_metrics = await db.execute(stmt)
    metrics_map = {row.org_id: row for row in branch_metrics.fetchall()}
    
    # Active queues
    q_stmt = select(Queue.org_id, func.count(Queue.id).label("count")).where(
        Queue.org_id.in_(org_ids), Queue.is_active == True, Queue.is_deleted == False
    ).group_by(Queue.org_id)
    q_res = await db.execute(q_stmt)
    q_counts = {row.org_id: row.count for row in q_res.fetchall()}
    
    # Sessions
    s_stmt = select(Session.org_id, func.count(Session.id).label("count")).where(Session.org_id.in_(org_ids))
    if start_date:
        s_stmt = s_stmt.where(Session.session_date >= start_date)
    if end_date:
        s_stmt = s_stmt.where(Session.session_date <= end_date)
    s_stmt = s_stmt.group_by(Session.org_id)
    s_res = await db.execute(s_stmt)
    s_counts = {row.org_id: row.count for row in s_res.fetchall()}
    
    # Staff
    staff_stmt = select(
        User.org_id,
        func.count(User.id).label("total"),
        func.sum(func.cast(User.is_active, func.integer())).label("active")
    ).where(User.org_id.in_(org_ids), User.role.in_(["admin", "staff"])).group_by(User.org_id)
    staff_res = await db.execute(staff_stmt)
    staff_counts = {row.org_id: {"total": row.total, "active": row.active} for row in staff_res.fetchall()}
    
    data = []
    for org in orgs:
        m = metrics_map.get(org.id)
        sc = staff_counts.get(org.id, {"total": 0, "active": 0})
        total = m.total if m else 0
        served = m.served if m else 0
        rate = round((served / total) * 100, 1) if total else 0
        avg_w = m.avg_wait if m else 0
        active_q = q_counts.get(org.id, 0)
        act_staff = sc["active"] or 1
        
        score = (rate * 0.5) + ((100 - min(avg_w/60, 100)) * 0.3) + (min(total/100, 1) * 20)
        
        data.append({
            "Branch Name": org.name,
            "Total Customers": total,
            "Customers Served": served,
            "Waiting": m.waiting if m else 0,
            "Cancelled": m.cancelled if m else 0,
            "Average Waiting Time": _format_time(avg_w),
            "Maximum Waiting Time": _format_time(m.max_wait if m else 0),
            "Average Service Time": _format_time(m.avg_serve if m else 0),
            "Maximum Service Time": _format_time(m.max_serve if m else 0),
            "Service Completion %": f"{rate}%",
            "Total Staff": sc["total"] or 0,
            "Active Staff": sc["active"] or 0,
            "Active Queues": active_q,
            "Active Sessions": s_counts.get(org.id, 0),
            "Customers per Staff": round(total / act_staff, 1),
            "Customers per Active Queue": round(total / active_q, 1) if active_q else total,
            "_score": score,
            "_raw_wait": avg_w,
            "_raw_rate": rate
        })
        
    data.sort(key=lambda x: x["_score"], reverse=True)
    
    summary = []
    if data:
        summary.append({"Ranking": "🥇 Best Overall Branch", "Branch Name": data[0]["Branch Name"], "Score Detail": f"{data[0]['Service Completion %']} completion"})
    if len(data) > 1:
        summary.append({"Ranking": "🥈 Second Best", "Branch Name": data[1]["Branch Name"], "Score Detail": f"{data[1]['Service Completion %']} completion"})
    if len(data) > 2:
        summary.append({"Ranking": "🥉 Third Best", "Branch Name": data[2]["Branch Name"], "Score Detail": f"{data[2]['Service Completion %']} completion"})
        
    most_active = max(data, key=lambda x: x["Total Customers"]) if data else None
    if most_active:
        summary.append({"Ranking": "🔥 Most Active Branch", "Branch Name": most_active["Branch Name"], "Score Detail": f"{most_active['Total Customers']} customers"})
        
    needs_improve = min(data, key=lambda x: x["_score"]) if data else None
    if needs_improve:
        summary.append({"Ranking": "⚠️ Needs Improvement", "Branch Name": needs_improve["Branch Name"], "Score Detail": f"{needs_improve['Service Completion %']} completion, {_format_time(needs_improve['_raw_wait'])} avg wait"})

    for d in data:
        d.pop("_score")
        d.pop("_raw_wait")
        d.pop("_raw_rate")
        
    sections = [
        {"title": "Overall Performance Rankings", "data": summary},
        {"title": "Branch Performance Details", "data": data}
    ]
    
    if format_type == "CSV":
        await _generate_csv(sections, file_path)
    else:
        await _generate_structured_excel(sections, file_path)

async def _report_staff_performance(job: ExportJob, db: AsyncSession, file_path: str, format_type: str):
    start_date, end_date, tz = await _resolve_date_bounds(db, job)
    org_ids, org_name_map, orgs = await _get_org_ids(db, job.parent_org_id, job.filters)
    
    if not org_ids:
        return [{"Message": "No branches found."}]
        
    from app.models.token import TokenStatus
    from sqlalchemy import or_
    
    stmt = select(
        User.id,
        User.first_name, User.last_name, User.email, User.role, User.last_active_at, User.org_id,
        func.count(Token.id).filter(Token.status == TokenStatus.done).label("served"),
        func.count(Token.id).filter(Token.status == TokenStatus.done, Token.completed_by_id == User.id).label("completed"),
        func.avg(func.extract('epoch', Token.completed_at - Token.served_at)).filter(Token.status == TokenStatus.done).label("avg_serve"),
        func.max(func.extract('epoch', Token.completed_at - Token.served_at)).filter(Token.status == TokenStatus.done).label("max_serve"),
        func.avg(func.extract('epoch', Token.served_at - Token.created_at)).filter(Token.status == TokenStatus.done).label("avg_wait")
    ).outerjoin(
        Token, and_(Token.served_by_id == User.id, Token.org_id == User.org_id)
    )
    
    if start_date:
        start_dt = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=tz)
        stmt = stmt.where(or_(Token.created_at >= start_dt, Token.id == None))
    if end_date:
        end_dt = datetime.combine(end_date, datetime.max.time()).replace(tzinfo=tz)
        stmt = stmt.where(or_(Token.created_at <= end_dt, Token.id == None))

    stmt = stmt.where(User.org_id.in_(org_ids), User.role.in_(["admin", "staff"]))
    stmt = stmt.group_by(User.id)
    
    res = await db.execute(stmt)
    staff_data = res.fetchall()
    
    days = 1
    if start_date and end_date:
        days = (end_date - start_date).days + 1
        
    two_mins_ago = datetime.now(timezone.utc) - timedelta(minutes=2)
    
    data = []
    for r in staff_data:
        name = f"{r.first_name or ''} {r.last_name or ''}".strip() or "—"
        online = "Online" if r.last_active_at and r.last_active_at.replace(tzinfo=timezone.utc) >= two_mins_ago else "Offline"
        
        serve_rate = round((r.completed / r.served) * 100, 1) if r.served else 0
        daily_avg = round(r.served / days, 1) if days > 0 else r.served
        
        # Avoid division by zero
        safe_served = max(r.served, 1)
        score = (r.served * 2) + (serve_rate * 0.5) + (max(600 - (r.avg_serve or 600), 0) / 60)
        
        data.append({
            "Name": name,
            "Email": r.email,
            "Branch": org_name_map.get(r.org_id, "Unknown"),
            "Role": r.role.capitalize(),
            "Status": online,
            "Last Active": r.last_active_at.strftime("%Y-%m-%d %H:%M") if r.last_active_at else "Never",
            "Customers Served": r.served,
            "Customers Completed": r.completed,
            "Service Completion %": f"{serve_rate}%",
            "Avg Customers/Day": daily_avg,
            "Avg Service Time": _format_time(r.avg_serve or 0),
            "Max Service Time": _format_time(r.max_serve or 0),
            "Avg Wait of Handled Customers": _format_time(r.avg_wait or 0),
            "_score": score
        })
        
    data.sort(key=lambda x: x["_score"], reverse=True)
    
    summary = []
    if data:
        best = data[0]
        summary.append({"Metric": "🌟 Best Performer", "Staff Member": best["Name"], "Detail": f"{best['Customers Served']} served"})
        
        highest_wl = max(data, key=lambda x: x["Customers Served"])
        summary.append({"Metric": "📈 Highest Workload", "Staff Member": highest_wl["Name"], "Detail": f"{highest_wl['Customers Served']} served"})
        
        active_staff = [d for d in data if d["Customers Served"] > 0]
        if active_staff:
            lowest_wl = min(active_staff, key=lambda x: x["Customers Served"])
            summary.append({"Metric": "📉 Lowest Workload (Active)", "Staff Member": lowest_wl["Name"], "Detail": f"{lowest_wl['Customers Served']} served"})

    for d in data:
        d.pop("_score")

    sections = [
        {"title": "Staff Summary", "data": summary},
        {"title": "Staff Performance Details", "data": data}
    ]
    
    if format_type == "CSV":
        await _generate_csv(sections, file_path)
    else:
        await _generate_structured_excel(sections, file_path)

async def _report_waiting_time_analysis(job: ExportJob, db: AsyncSession, file_path: str, format_type: str):
    start_date, end_date, tz = await _resolve_date_bounds(db, job)
    org_ids, org_name_map, orgs = await _get_org_ids(db, job.parent_org_id, job.filters)
    
    if not org_ids:
        return [{"Message": "No branches found."}]
        
    from app.models.token import TokenStatus
    
    stmt = select(
        Token.org_id,
        func.count(Token.id).label("served"),
        func.avg(func.extract('epoch', Token.served_at - Token.created_at)).label("avg_wait"),
        func.max(func.extract('epoch', Token.served_at - Token.created_at)).label("max_wait"),
        func.min(func.extract('epoch', Token.served_at - Token.created_at)).label("min_wait")
    ).where(Token.org_id.in_(org_ids), Token.status == TokenStatus.done)
    stmt = _apply_date_filter(stmt, start_date, end_date, tz)
    stmt = stmt.group_by(Token.org_id)
    
    res = await db.execute(stmt)
    branch_metrics = res.fetchall()
    
    tot_served = sum(r.served for r in branch_metrics)
    avg_wait = sum((r.avg_wait or 0) * r.served for r in branch_metrics) / tot_served if tot_served else 0
    max_wait = max((r.max_wait for r in branch_metrics if r.max_wait), default=0)
    min_wait = min((r.min_wait for r in branch_metrics if r.min_wait), default=0) if tot_served else 0
    
    overall = [
        {"Average Waiting Time": _format_time(avg_wait), "Maximum Waiting Time": _format_time(max_wait), "Minimum Waiting Time": _format_time(min_wait)}
    ]
    
    # Distribution
    dist_stmt = select(
        func.count().filter((func.extract('epoch', Token.served_at - Token.created_at) / 60) < 5).label("0-5"),
        func.count().filter(and_((func.extract('epoch', Token.served_at - Token.created_at) / 60) >= 5, (func.extract('epoch', Token.served_at - Token.created_at) / 60) < 10)).label("5-10"),
        func.count().filter(and_((func.extract('epoch', Token.served_at - Token.created_at) / 60) >= 10, (func.extract('epoch', Token.served_at - Token.created_at) / 60) < 20)).label("10-20"),
        func.count().filter(and_((func.extract('epoch', Token.served_at - Token.created_at) / 60) >= 20, (func.extract('epoch', Token.served_at - Token.created_at) / 60) < 30)).label("20-30"),
        func.count().filter((func.extract('epoch', Token.served_at - Token.created_at) / 60) >= 30).label("30+")
    ).where(Token.org_id.in_(org_ids), Token.status == TokenStatus.done)
    dist_stmt = _apply_date_filter(dist_stmt, start_date, end_date, tz)
    
    dist_res = await db.execute(dist_stmt)
    dist = dist_res.first()
    
    distribution = []
    if dist:
        total_d = tot_served or 1
        distribution = [
            {"Wait Time": "0-5 min", "Customers": dist[0], "% of Total": f"{round((dist[0]/total_d)*100,1)}%"},
            {"Wait Time": "5-10 min", "Customers": dist[1], "% of Total": f"{round((dist[1]/total_d)*100,1)}%"},
            {"Wait Time": "10-20 min", "Customers": dist[2], "% of Total": f"{round((dist[2]/total_d)*100,1)}%"},
            {"Wait Time": "20-30 min", "Customers": dist[3], "% of Total": f"{round((dist[3]/total_d)*100,1)}%"},
            {"Wait Time": "30+ min", "Customers": dist[4], "% of Total": f"{round((dist[4]/total_d)*100,1)}%"}
        ]
        
    branch_comp = []
    for r in branch_metrics:
        branch_comp.append({
            "Branch Name": org_name_map.get(r.org_id, "Unknown"),
            "Customers Served": r.served,
            "Average Wait": _format_time(r.avg_wait or 0),
            "Maximum Wait": _format_time(r.max_wait or 0),
            "_raw_wait": r.avg_wait or 0
        })
    branch_comp.sort(key=lambda x: x["_raw_wait"])
    
    for idx, b in enumerate(branch_comp, 1):
        b["Rank (Lowest Wait)"] = f"#{idx}"
        b.pop("_raw_wait")
        
    # Peak Analysis
    # Ensure correct time extraction according to timezone
    try:
        hour_expr = func.extract('hour', func.timezone(tz.key, Token.created_at))
        day_expr = func.date(func.timezone(tz.key, Token.created_at))
    except Exception:
        hour_expr = func.extract('hour', Token.created_at)
        day_expr = func.date(Token.created_at)
        
    peak_stmt = select(
        hour_expr.label("hr"),
        func.count(Token.id).label("count")
    ).where(Token.org_id.in_(org_ids), Token.status != TokenStatus.deleted)
    peak_stmt = _apply_date_filter(peak_stmt, start_date, end_date, tz)
    peak_stmt = peak_stmt.group_by("hr")
    
    peak_res = await db.execute(peak_stmt)
    hours = peak_res.fetchall()
    
    day_stmt = select(
        day_expr.label("dy"),
        func.count(Token.id).label("count")
    ).where(Token.org_id.in_(org_ids), Token.status != TokenStatus.deleted)
    day_stmt = _apply_date_filter(day_stmt, start_date, end_date, tz)
    day_stmt = day_stmt.group_by("dy")
    
    day_res = await db.execute(day_stmt)
    days = day_res.fetchall()
    
    peak_hour = max(hours, key=lambda x: x.count) if hours else None
    busy_day = max(days, key=lambda x: x.count) if days else None
    
    peak_analysis = []
    if peak_hour:
        h = int(peak_hour.hr)
        h_str = f"{h % 12 or 12}:00 {'AM' if h < 12 else 'PM'}"
        peak_analysis.append({"Metric": "Peak Hour", "Value": h_str, "Customers": peak_hour.count})
    if busy_day:
        peak_analysis.append({"Metric": "Busiest Day", "Value": str(busy_day.dy), "Customers": busy_day.count})
        
    insights = []
    if branch_comp:
        insights.append({"Observation": "Lowest Wait Time Branch", "Branch": branch_comp[0]["Branch Name"]})
        insights.append({"Observation": "Highest Wait Time Branch", "Branch": branch_comp[-1]["Branch Name"]})

    sections = [
        {"title": "Overall Waiting Metrics", "data": overall},
        {"title": "Waiting Time Distribution", "data": distribution},
        {"title": "Peak Analysis", "data": peak_analysis},
        {"title": "Key Insights", "data": insights},
        {"title": "Branch Comparison", "data": branch_comp}
    ]
    
    if format_type == "CSV":
        await _generate_csv(sections, file_path)
    else:
        await _generate_structured_excel(sections, file_path)


async def generate_export(job_id: uuid.UUID):
    from app.db.session import AsyncSessionLocal
    async with AsyncSessionLocal() as db:
        job = await db.get(ExportJob, job_id)
    if not job:
        return
        
    try:
        job.status = "processing"
        await db.commit()
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{job.report_type.replace(' ', '_')}_{timestamp}"
        
        if job.report_type == "Customer Detailed Report":
            file_path = await _generate_customer_detailed_report(job, db, filename)
        else:
            fmt = job.format.upper()
            if fmt == "PDF":
                fmt = "EXCEL" # Replaced custom pdf generation with well-formatted single sheet excel
                
            file_path = os.path.join(EXPORTS_DIR, f"{filename}.{'csv' if fmt == 'CSV' else 'xlsx'}")
            
            if job.report_type == "Executive Summary":
                await _report_executive_summary(job, db, file_path, fmt)
            elif job.report_type == "Branch Performance Report":
                await _report_branch_performance(job, db, file_path, fmt)
            elif job.report_type == "Staff Performance Report":
                await _report_staff_performance(job, db, file_path, fmt)
            elif job.report_type == "Waiting Time Analysis":
                await _report_waiting_time_analysis(job, db, file_path, fmt)
            else:
                raise ValueError(f"Unknown report type: {job.report_type}")

        job.file_path = file_path
        job.status = "completed"
        job.completed_at = datetime.now(timezone.utc)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        job.status = "failed"
        job.error_message = str(e)
        job.completed_at = datetime.now(timezone.utc)
        
    finally:
        await db.commit()

def _generate_pdf(df, file_path, title):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    doc = SimpleDocTemplate(file_path, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    elements.append(Paragraph(title, styles['Title']))
    
    data = [df.columns[:,].values.astype(str).tolist()] + df.values.tolist()
    
    t = Table(data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ]))
    
    elements.append(t)
    doc.build(elements)

# ── Customer Detailed Report Logic ──

async def _generate_customer_detailed_report(job: ExportJob, db: AsyncSession, filename: str) -> str:
    # Build Query
    date_col = func.coalesce(Session.session_date, func.date(Token.created_at))
    
    from sqlalchemy.orm import aliased
    CompletedByUser = aliased(User)

    stmt = (
        select(
            Organization.name.label("branch_name"),
            date_col.label("date"),
            Queue.name.label("queue_name"),
            Token.token_number,
            Token.assigned_line,
            Token.customer_name,
            Token.customer_phone,
            Token.companion_names,
            Token.status,
            Token.created_at,
            Token.served_at,
            Token.completed_at,
            Token.called_via_invite,
            Token.served_by_id,
            Token.completed_by_id,
            Token.entry_type,
            Session.title.label("session_name"),
            User.first_name.label("staff_first"),
            User.last_name.label("staff_last"),
            User.email.label("staff_email"),
            CompletedByUser.first_name.label("completed_first"),
            CompletedByUser.last_name.label("completed_last"),
            CompletedByUser.email.label("completed_email"),
            Token.removed_by,
            Token.skipped_at,
            Token.recalled_at,
            Token.deleted_at
        )
        .select_from(Token)
        .join(Organization, Token.org_id == Organization.id)
        .outerjoin(Session, Token.session_id == Session.id)
        .join(Queue, Token.queue_id == Queue.id)
        .outerjoin(User, Token.served_by_id == User.id)
        .outerjoin(CompletedByUser, Token.completed_by_id == CompletedByUser.id)
        .where(Organization.parent_organization_id == job.parent_org_id)
    )

    filters = job.filters or {}
    
    # 1. Branch filter
    branch_ids = filters.get("branch_ids")
    if branch_ids:
        stmt = stmt.where(Organization.id.in_([uuid.UUID(b) for b in branch_ids]))
        
    # 2. Queue filter
    queue_names = filters.get("queue_names")
    if queue_names:
        stmt = stmt.where(Queue.name.in_(queue_names))
        
    # 3. Date Range
    date_range = filters.get("date_range", "All Time")
    now = datetime.now(timezone.utc)
    today = now.date()
    
    if date_range == "Today":
        stmt = stmt.where(date_col == today)
    elif date_range == "Yesterday":
        stmt = stmt.where(date_col == today - timedelta(days=1))
    elif date_range == "Last 7 Days":
        stmt = stmt.where(date_col >= today - timedelta(days=7))
    elif date_range == "Last 30 Days":
        stmt = stmt.where(date_col >= today - timedelta(days=30))
    elif date_range == "This Month":
        start_of_month = today.replace(day=1)
        stmt = stmt.where(date_col >= start_of_month)
    elif date_range == "Last Month":
        start_of_this_month = today.replace(day=1)
        end_of_last_month = start_of_this_month - timedelta(days=1)
        start_of_last_month = end_of_last_month.replace(day=1)
        stmt = stmt.where(date_col.between(start_of_last_month, end_of_last_month))
    elif date_range == "Custom Date Range":
        c_start = filters.get("custom_start_date")
        c_end = filters.get("custom_end_date")
        if c_start:
            stmt = stmt.where(date_col >= datetime.strptime(c_start, "%Y-%m-%d").date())
        if c_end:
            stmt = stmt.where(date_col <= datetime.strptime(c_end, "%Y-%m-%d").date())

    # Order by hierarchy
    stmt = stmt.order_by(
        Organization.name,
        date_col.desc(),
        Queue.name,
        Token.created_at
    )

    res = await db.execute(stmt)
    records = res.all()

    # Process data
    formatted_data = []
    total_wait_secs = 0
    total_serve_secs = 0
    served_count = 0
    cancelled_count = 0

    unique_branches = set()
    unique_queues = set()

    for r in records:
        branch = r.branch_name
        q_name = r.queue_name
        unique_branches.add(branch)
        unique_queues.add(f"{branch}-{q_name}")
        
        # Calculate Wait Time
        wait_mins = None
        if r.served_at and r.created_at:
            w_secs = (r.served_at - r.created_at).total_seconds()
            wait_mins = round(w_secs / 60, 2)
            if r.status == "done":
                total_wait_secs += w_secs
                
        # Calculate Service Time
        serve_mins = None
        if r.completed_at and r.served_at:
            s_secs = (r.completed_at - r.served_at).total_seconds()
            serve_mins = round(s_secs / 60, 2)
            if r.status == "done":
                total_serve_secs += s_secs
                
        if r.status == "done":
            served_count += 1
        elif r.status in ["skipped", "deleted"]:
            cancelled_count += 1

        staff_name = ""
        if r.staff_first or r.staff_last:
            staff_name = f"{r.staff_first or ''} {r.staff_last or ''}".strip()
        elif r.staff_email:
            staff_name = r.staff_email.split('@')[0]
            
        completed_by_name = ""
        if r.completed_first or r.completed_last:
            completed_by_name = f"{r.completed_first or ''} {r.completed_last or ''}".strip()
        elif r.completed_email:
            completed_by_name = r.completed_email.split('@')[0]


        # Companions: stored as JSON list of names
        companions_raw = r.companion_names or []
        companions_str = ", ".join(companions_raw) if companions_raw else ""

        # Entry Type
        if r.entry_type == "manual":
            entry_type_label = "Manual"
        elif r.entry_type == "qr":
            entry_type_label = "QR Code"
        elif r.entry_type == "auto":
            entry_type_label = "Auto"
        else:
            entry_type_label = r.entry_type or ""

        # Format Removed By
        removed_by_label = ""
        if r.removed_by == "customer":
            removed_by_label = "Customer"
        elif r.removed_by == "session_end":
            removed_by_label = "System (Session End)"
        elif r.removed_by:
            removed_by_label = "Staff"

        formatted_data.append({
            "Branch Name": branch,
            "Date": r.date.strftime("%Y-%m-%d") if r.date else "",
            "Queue Name": q_name,
            "Token Number": r.token_number,
            "Service Line": r.assigned_line if r.assigned_line is not None else "",
            "Customer Name": r.customer_name or "",
            "Customer Phone": r.customer_phone or "",
            "Companions": companions_str,
            "Status": r.status.title(),
            "Created At": r.created_at.strftime("%H:%M:%S") if r.created_at else "",
            "Served At": r.served_at.strftime("%H:%M:%S") if r.served_at else "",
            "Completed At": r.completed_at.strftime("%H:%M:%S") if r.completed_at else "",
            "Skipped At": r.skipped_at.strftime("%H:%M:%S") if r.skipped_at else "",
            "Recalled At": r.recalled_at.strftime("%H:%M:%S") if r.recalled_at else "",
            "Removed At": r.deleted_at.strftime("%H:%M:%S") if r.deleted_at else "",
            "Wait Time (mins)": wait_mins,
            "Serve Time (mins)": serve_mins,
            "Served By": staff_name,
            "Completed By": completed_by_name,
            "Removed By": removed_by_label,
            "Call Method": "Skipped" if r.status == "skipped" else "Normal",
            "Entry Type": entry_type_label
        })

    # Add Called At back as empty
    for row in formatted_data:
        # insert Called At after Created At
        # Since dictionaries are ordered in Python 3.7+, we'll rebuild the dict to ensure column order
        pass

    # Rebuild data with exact columns in the required order
    final_data = []
    for r in formatted_data:
        final_data.append({
            "Branch Name": r.get("Branch Name", ""),   # Keep for grouping/sheet splitting
            "Queue Name": r.get("Queue Name", ""),     # Keep for grouping/sheet splitting
            "Date": r.get("Date", ""),
            "Token Number": r.get("Token Number", ""),
            "Queue": r.get("Queue Name", ""),
            "Service Line": r.get("Service Line", ""),
            "Customer Name": r.get("Customer Name", ""),
            "Customer Phone": r.get("Customer Phone", ""),
            "Companions": r.get("Companions", ""),
            "Status": r.get("Status", ""),
            "Created At": r.get("Created At", ""),
            "Served At": r.get("Served At", ""),
            "Completed At": r.get("Completed At", ""),
            "Skipped At": r.get("Skipped At", ""),
            "Recalled At": r.get("Recalled At", ""),
            "Removed At": r.get("Removed At", ""),
            "Wait Time (mins)": r.get("Wait Time (mins)", ""),
            "Serve Time (mins)": r.get("Serve Time (mins)", ""),
            "Served By": r.get("Served By", ""),
            "Completed By": r.get("Completed By", ""),
            "Removed By": r.get("Removed By", ""),
            "Call Method": r.get("Call Method", ""),
            "Entry Type": r.get("Entry Type", "")
        })

    df = pd.DataFrame(final_data)

    # Compute Summary
    total_customers = len(formatted_data)
    avg_wait = round(total_wait_secs / served_count / 60, 2) if served_count > 0 else 0
    avg_serve = round(total_serve_secs / served_count / 60, 2) if served_count > 0 else 0

    summary_df = pd.DataFrame([
        {"Metric": "Total Branches", "Value": len(unique_branches)},
        {"Metric": "Total Queues", "Value": len(unique_queues)},
        {"Metric": "Total Customers", "Value": total_customers},
        {"Metric": "Total Served", "Value": served_count},
        {"Metric": "Total Cancelled / Deleted", "Value": cancelled_count},
        {"Metric": "Average Wait Time (Mins)", "Value": avg_wait},
        {"Metric": "Average Service Time (Mins)", "Value": avg_serve},
    ])

    # Fetch Parent Org Name
    from app.models.parent_organization import ParentOrganization
    parent_org = await db.scalar(select(ParentOrganization).where(ParentOrganization.id == job.parent_org_id))
    parent_org_name = parent_org.name if parent_org else "Parent Organization"
    report_title = f"{parent_org_name} Customer Report"

    if df.empty:
        # Fallback for empty data
        if job.format.upper() == "EXCEL":
            file_path = os.path.join(EXPORTS_DIR, f"{filename}.xlsx")
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Write an empty dataframe so we can grab the sheet and edit it
                pd.DataFrame().to_excel(writer, sheet_name='Report', index=False)
                wb = writer.book
                ws = wb.active
                
                # Title
                ws.append([report_title])
                ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=16)
                ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
                ws.append([])
                
                # Summary
                ws.append(["Report Summary"])
                ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=14)
                ws.append(["Metric", "Value"])
                ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
                ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
                
                for _, row in summary_df.iterrows():
                    ws.append([row['Metric'], row['Value']])
            return file_path
        elif job.format.upper() == "CSV":
            file_path = os.path.join(EXPORTS_DIR, f"{filename}.csv")
            with open(file_path, 'w') as f:
                summary_df.to_csv(f, index=False)
            return file_path
        else:
            file_path = os.path.join(EXPORTS_DIR, f"{filename}.pdf")
            _generate_detailed_pdf(df, summary_df, file_path, report_title)
            return file_path

    # Grouping logic
    cols_to_drop = ['Branch Name', 'Queue Name']

    if job.format.upper() == "EXCEL":
        file_path = os.path.join(EXPORTS_DIR, f"{filename}.xlsx")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Write Title
        ws.append([report_title])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=16)
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=6)
        ws.append([])
        
        # Write Summary
        ws.append(["Report Summary"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=14)
        ws.append(["Metric", "Value"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
        
        for _, row in summary_df.iterrows():
            ws.append([row['Metric'], row['Value']])
            
        ws.append([])
        
        # Nested hierarchical grouping: Branch -> Date -> Queue
        branches = df['Branch Name'].unique()
        for branch in branches:
            branch_df = df[df['Branch Name'] == branch]
            
            ws.append([f"BRANCH: {branch}"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=14)
            ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=6)
            ws.append([])
            ws.append(["---"])
            
            dates = branch_df['Date'].unique()
            for date in dates:
                date_df = branch_df[branch_df['Date'] == date]
                
                ws.append([f"DATE: {date}"])
                ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
                ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=6)
                ws.append([])
                ws.append(["---"])
                
                queues = date_df['Queue Name'].unique()
                for queue in queues:
                    queue_df = date_df[date_df['Queue Name'] == queue]
                    
                    ws.append([f"QUEUE: {queue}"])
                    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
                    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=6)
                    ws.append([])
                    
                    group_clean = queue_df.drop(columns=cols_to_drop, errors='ignore')
                    
                    # Table headers
                    row_idx = ws.max_row + 1
                    ws.append(list(group_clean.columns))
                    for col_idx in range(1, len(group_clean.columns) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                        cell.border = thin_border
                    
                    for _, row in group_clean.iterrows():
                        ws.append(list(row))
                        # apply border to data rows
                        r_idx = ws.max_row
                        for c_idx in range(1, len(group_clean.columns) + 1):
                            ws.cell(row=r_idx, column=c_idx).border = thin_border
                        
                    ws.append([]) # Empty row for spacing
                    ws.append(["---"])
                    ws.append([])
            
        # Adjust column widths for readability
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if type(cell).__name__ != 'MergedCell':
                    try:
                        val = str(cell.value)
                        if cell.value is not None and len(val) > max_length:
                            max_length = len(val)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 50:
                adjusted_width = 50 # Cap maximum width
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(file_path)
        return file_path
        
    elif job.format.upper() == "CSV":
        file_path = os.path.join(EXPORTS_DIR, f"{filename}.csv")
        with open(file_path, 'w') as f:
            f.write(f"# {report_title}\n\n")
            f.write("--- REPORT SUMMARY ---\n")
            summary_df.to_csv(f, index=False)
            f.write("\n\n--- DETAILED RECORDS ---\n")
            
            grouped = df.groupby(['Branch Name', 'Date', 'Queue Name'])
            for (branch, date, queue), group in grouped:
                f.write(f"\n## Branch: {branch}\n")
                f.write(f"### Date: {date}\n")
                f.write(f"#### Queue: {queue}\n")
                
                group_clean = group.drop(columns=cols_to_drop)
                group_clean.to_csv(f, index=False)
                
        return file_path
        
    elif job.format.upper() == "PDF":
        file_path = os.path.join(EXPORTS_DIR, f"{filename}.pdf")
        _generate_detailed_pdf(df, summary_df, file_path, report_title)
        return file_path
        
    else:
        raise ValueError(f"Unsupported format {job.format}")

def _generate_detailed_pdf(df, summary_df, file_path, title):
    from reportlab.lib.pagesizes import landscape, A3
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    doc = SimpleDocTemplate(file_path, pagesize=landscape(A3), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 20))
    
    # Summary Table
    elements.append(Paragraph("Report Summary", styles['Heading2']))
    sum_data = [summary_df.columns.values.astype(str).tolist()] + summary_df.values.tolist()
    sum_t = Table(sum_data)
    sum_t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 10),
        ('BACKGROUND', (0,1), (-1,-1), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 1, colors.lightgrey)
    ]))
    elements.append(sum_t)
    elements.append(Spacer(1, 30))
    
    if df.empty:
        elements.append(Paragraph("No customer records found for the selected criteria.", styles['Normal']))
        doc.build(elements)
        return
        
    # Limit rows for PDF to avoid memory issues
    if len(df) > 1000:
        elements.append(Paragraph("Note: PDF output is limited to 1000 rows. Please export to Excel/CSV for the full dataset.", styles['Normal']))
        elements.append(Spacer(1, 10))
        df = df.head(1000)
        
    # Group hierarchical data
    cols_to_drop = ['Branch Name', 'Date', 'Queue Name']
    grouped = df.groupby(['Branch Name', 'Date', 'Queue Name'])
    
    for (branch, date, queue), group in grouped:
        elements.append(Paragraph(f"Branch: {branch}", styles['Heading3']))
        elements.append(Paragraph(f"Date: {date}", styles['Normal']))
        elements.append(Paragraph(f"Queue: {queue}", styles['Normal']))
        elements.append(Spacer(1, 10))
        
        group_clean = group.drop(columns=cols_to_drop)
        data = [group_clean.columns.values.astype(str).tolist()] + group_clean.astype(str).values.tolist()
        
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#6b7280')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('BACKGROUND', (0,1), (-1,-1), colors.white),
            ('GRID', (0,0), (-1,-1), 1, colors.lightgrey)
        ]))
        elements.append(t)
        elements.append(Spacer(1, 20))
        
    doc.build(elements)
