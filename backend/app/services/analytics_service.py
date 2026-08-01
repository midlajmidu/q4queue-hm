"""
app/services/analytics_service.py
Service for fetching overview statistics and graphs.
"""
import uuid
from typing import Optional

from sqlalchemy import and_, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.token import Token, TokenStatus

async def get_overview_metrics(
    db: AsyncSession,
    *,
    org_id: uuid.UUID,
    session_id: Optional[uuid.UUID] = None,
    queue_id: Optional[uuid.UUID] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    recent_limit: int = 5,
    recent_offset: int = 0,
) -> dict:
    """Fetch aggregated metrics for the dashboard."""
    from dateutil.parser import parse as parse_date
    from app.models.organization import Organization
    org = await db.scalar(select(Organization).where(Organization.id == org_id))
    org_tz_str = org.timezone if org and org.timezone else "UTC"
    
    # Base conditions
    conditions = [Token.org_id == org_id]
    
    from app.models.queue import Queue
    
    if queue_id:
        conditions.append(Token.queue_id == queue_id)
        
    if start_date:
        try:
            from zoneinfo import ZoneInfo
            tz = ZoneInfo(org_tz_str)
            dt = parse_date(start_date)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=tz)
            else:
                dt = dt.astimezone(tz)
            conditions.append(Token.created_at >= dt.astimezone(ZoneInfo("UTC")))
        except Exception:
            pass
            
    if end_date:
        try:
            from zoneinfo import ZoneInfo
            tz = ZoneInfo(org_tz_str)
            ed = parse_date(end_date)
            if ed.tzinfo is None:
                ed = ed.replace(tzinfo=tz)
            else:
                ed = ed.astimezone(tz)
            if ed.hour == 0 and ed.minute == 0 and ed.second == 0:
                ed = ed.replace(hour=23, minute=59, second=59, microsecond=999999)
            conditions.append(Token.created_at <= ed.astimezone(ZoneInfo("UTC")))
        except Exception:
            pass
    
    join_queue = False
    if session_id:
        from app.models.session import Session
        sess = await db.scalar(select(Session).where(Session.id == session_id))
        if sess:
            conditions.append(Token.queue_id == sess.queue_id)
            conditions.append(func.date(func.timezone(org_tz_str, Token.created_at)) == sess.session_date)
        else:
            conditions.append(Token.id == None)

    # Filter out deleted tokens from most metrics
    active_conditions = conditions.copy()
    from app.models.token import TokenStatus
    active_conditions.append(Token.status != TokenStatus.deleted)

    # 1. Status Counts
    count_query = select(Token.status, func.count(Token.id)).where(and_(*conditions)).group_by(Token.status)
    if join_queue:
        count_query = count_query.join(Queue, Token.queue_id == Queue.id)
    
    count_result = await db.execute(count_query)
    
    counts = {s.value: 0 for s in TokenStatus}
    for row in count_result.all():
        counts[row[0].value] = row[1]
        
    # Total visits includes all tokens (even deleted ones) for accurate dashboard math
    total_visits = counts[TokenStatus.waiting.value] + counts[TokenStatus.serving.value] + \
                   counts[TokenStatus.done.value] + counts[TokenStatus.skipped.value] + \
                   counts[TokenStatus.deleted.value]
    
    served_visits = counts[TokenStatus.done.value]
    cancelled_visits = counts[TokenStatus.skipped.value] + counts[TokenStatus.deleted.value]
    waiting_visits = counts[TokenStatus.waiting.value]

    # Calculate invited tokens
    invited_query = select(func.count(Token.id)).where(and_(*active_conditions, Token.called_via_invite == True))
    if join_queue:
        invited_query = invited_query.join(Queue, Token.queue_id == Queue.id)
    invited_res = await db.execute(invited_query)
    invited_visits = invited_res.scalar_one()

    # 2. Timing Aggregations - Exclude deleted
    timing_query = select(
        func.avg(func.extract('epoch', Token.served_at - Token.created_at)).label('avg_wait_sec'),
        func.max(func.extract('epoch', Token.served_at - Token.created_at)).label('max_wait_sec'),
        func.avg(func.extract('epoch', Token.completed_at - Token.served_at)).label('avg_serve_sec'),
        func.max(func.extract('epoch', Token.completed_at - Token.served_at)).label('max_serve_sec'),
    ).where(and_(*active_conditions))
    
    if join_queue:
        timing_query = timing_query.join(Queue, Token.queue_id == Queue.id)

    timing_res = await db.execute(timing_query)
    row = timing_res.first()
    
    def format_time(seconds: float | None) -> str:
        if not seconds:
            return "00:00:00"
        m, s = divmod(int(seconds), 60)
        h, m = divmod(m, 60)
        return f"{h:02d}:{m:02d}:{s:02d}"

    # 3. Hourly Chart (Visits by hour) - Exclude deleted
    hourly_query = select(
        func.extract('hour', func.timezone(org_tz_str, Token.created_at)).label('hr'),
        func.count(Token.id)
    ).where(and_(*active_conditions)).group_by('hr').order_by('hr')
    
    if join_queue:
        hourly_query = hourly_query.join(Queue, Token.queue_id == Queue.id)

    hourly_res = await db.execute(hourly_query)
    hourly_data = [{"hour": f"{int(row[0]):02d}:00", "visits": row[1]} for row in hourly_res.all()]

    # 4. Monthly Chart - Exclude deleted
    monthly_query = select(
        func.extract('month', func.timezone(org_tz_str, Token.created_at)).label('mon'),
        func.extract('year', func.timezone(org_tz_str, Token.created_at)).label('yr'),
        func.count(Token.id)
    ).where(and_(*active_conditions)).group_by('yr', 'mon').order_by('yr', 'mon')
    
    if join_queue:
        monthly_query = monthly_query.join(Queue, Token.queue_id == Queue.id)

    monthly_res = await db.execute(monthly_query)
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    monthly_data = [{"month": f"{months[int(row[0])-1]} {int(row[1])}", "visits": row[2]} for row in monthly_res.all()]

    # 5. Daily Timings Chart - Exclude deleted
    daily_timings_query = select(
        func.date(func.timezone(org_tz_str, Token.created_at)).label('dt'),
        func.avg(func.extract('epoch', Token.served_at - Token.created_at)).label('avg_wait'),
        func.avg(func.extract('epoch', Token.completed_at - Token.served_at)).label('avg_serve'),
    ).where(and_(*active_conditions)).group_by('dt').order_by('dt')
    
    if join_queue:
        daily_timings_query = daily_timings_query.join(Queue, Token.queue_id == Queue.id)

    daily_timings_res = await db.execute(daily_timings_query)
    daily_timings_data = [
        {
            "date": row.dt.isoformat() if row.dt else "",
            "avg_wait": float(row.avg_wait) if row.avg_wait else 0,
            "avg_serve": float(row.avg_serve) if row.avg_serve else 0,
        }
        for row in daily_timings_res.all()
    ]

    # 6. Staff Performance - completed/served tokens by user
    from app.models.user import User
    staff_perf_query = select(
        User.id,
        User.first_name,
        User.last_name,
        User.email,
        func.count(Token.id).label('total_served'),
        func.avg(func.extract('epoch', Token.completed_at - Token.served_at)).label('avg_serve'),
    ).join(
        User, Token.served_by_id == User.id
    ).where(
        and_(*active_conditions, Token.status == TokenStatus.done)
    ).group_by(User.id).order_by(func.count(Token.id).desc())
    
    if join_queue:
        staff_perf_query = staff_perf_query.join(Queue, Token.queue_id == Queue.id)

    staff_perf_res = await db.execute(staff_perf_query)
    staff_performance_data = [
        {
            "staff_id": str(row.id),
            "name": f"{row.first_name} {row.last_name}".strip() if row.first_name else row.email.split('@')[0],
            "total_served": row.total_served,
            "avg_serve": float(row.avg_serve) if row.avg_serve else 0,
        }
        for row in staff_perf_res.all()
    ]

    from app.models.session import Session
    
    # 7. Recent Activity (for show last details request)
    recent_query = select(
        Token.token_number,
        Token.status,
        Token.created_at,
        Token.served_at,
        Token.completed_at,
        Token.skipped_at,
        Token.recalled_at,
        Token.customer_name,
        Queue.name.label('queue_name'),
        Queue.prefix.label('queue_prefix'),
        Session.title.label('session_title')
    ).join(Queue, Token.queue_id == Queue.id).outerjoin(
        Session, and_(Token.queue_id == Session.queue_id, func.date(func.timezone(org_tz_str, Token.created_at)) == Session.session_date)
    ).where(
        and_(*active_conditions)
    ).order_by(Token.created_at.desc()).limit(recent_limit).offset(recent_offset)
    
    # No special join needed for recent_activity as it already joins Queue

    recent_res = await db.execute(recent_query)
    recent_activity = [
        {
            "prefix": r.queue_prefix,
            "number": r.token_number,
            "status": r.status.value,
            "queue": r.queue_name,
            "session_name": r.session_title,
            "customer_name": r.customer_name or "Walk-in",
            "time": r.created_at.isoformat(),
            "served_at": r.served_at.isoformat() if r.served_at else None,
            "completed_at": r.completed_at.isoformat() if r.completed_at else None,
            "skipped_at": r.skipped_at.isoformat() if r.skipped_at else None,
            "recalled_at": r.recalled_at.isoformat() if r.recalled_at else None,
        }
        for r in recent_res.all()
    ]

    # 8. Longest waiting token for dynamic alerts
    longest_waiting_query = select(
        Queue.name.label('queue_name'),
        Session.title.label('session_title'),
        Session.session_date.label('session_date')
    ).select_from(Token).join(Queue, Token.queue_id == Queue.id).outerjoin(
        Session, and_(Token.queue_id == Session.queue_id, func.date(func.timezone(org_tz_str, Token.created_at)) == Session.session_date)
    ).where(
        and_(*active_conditions, Token.status == TokenStatus.waiting)
    ).order_by(Token.created_at.asc()).limit(1)

    longest_res = await db.execute(longest_waiting_query)
    longest_row = longest_res.first()
    
    longest_waiting_queue = None
    longest_waiting_session = None
    if longest_row:
        longest_waiting_queue = longest_row.queue_name
        longest_waiting_session = longest_row.session_title or str(longest_row.session_date)

    return {
        "status_counts": {
            "total": total_visits,
            "served": served_visits,
            "cancelled": cancelled_visits,
            "waiting": waiting_visits,
            "invited": invited_visits,
        },
        "timings": {
            "avg_waiting_time": format_time(row.avg_wait_sec),
            "max_waiting_time": format_time(row.max_wait_sec),
            "avg_served_time": format_time(row.avg_serve_sec),
            "max_served_time": format_time(row.max_serve_sec),
        },
        "charts": {
            "hourly": hourly_data,
            "monthly": monthly_data,
        },
        "daily_timings": daily_timings_data,
        "staff_performance": staff_performance_data,
        "recent_activity": recent_activity,
        "longest_waiting_queue": longest_waiting_queue,
        "longest_waiting_session": longest_waiting_session
    }

async def get_history_details(
    db: AsyncSession,
    *,
    org_id: uuid.UUID,
    session_id: Optional[uuid.UUID] = None,
    queue_id: Optional[uuid.UUID] = None,
    search: Optional[str] = None,
    status: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
) -> dict:
    """Fetch detailed token history with pagination and filters."""
    from app.models.queue import Queue
    from app.models.organization import Organization
    from sqlalchemy import or_, cast, String
    
    org = await db.scalar(select(Organization).where(Organization.id == org_id))
    org_tz_str = org.timezone if org and org.timezone else "UTC"
    
    conditions = [Token.org_id == org_id]
    if queue_id:
        conditions.append(Token.queue_id == queue_id)
    if status:
        conditions.append(Token.status == status)
    
    if search:
        search_term = f"%{search.lower()}%"
        conditions.append(or_(
            func.lower(Token.customer_name).like(search_term),
            Token.customer_phone.like(search_term),
            cast(Token.token_number, String).like(search_term)
        ))
    
    join_queue = False
    if session_id:
        from app.models.session import Session
        sess = await db.scalar(select(Session).where(Session.id == session_id))
        if sess:
            conditions.append(Token.queue_id == sess.queue_id)
            conditions.append(func.date(func.timezone(org_tz_str, Token.created_at)) == sess.session_date)
        else:
            conditions.append(Token.id == None)

    from sqlalchemy.orm import aliased
    from app.models.user import User
    
    ServedUser = aliased(User)
    CompletedUser = aliased(User)

    query = select(
        Token,
        Queue.name.label('queue_name'),
        Queue.prefix.label('queue_prefix'),
        ServedUser.first_name.label('served_first'),
        ServedUser.last_name.label('served_last'),
        ServedUser.email.label('served_email'),
        CompletedUser.first_name.label('completed_first'),
        CompletedUser.last_name.label('completed_last'),
        CompletedUser.email.label('completed_email')
    ).join(Queue, Token.queue_id == Queue.id).outerjoin(
        ServedUser, Token.served_by_id == ServedUser.id
    ).outerjoin(
        CompletedUser, Token.completed_by_id == CompletedUser.id
    ).where(
        and_(*conditions)
    ).order_by(Token.created_at.desc())

    # Count total for pagination
    count_query = select(func.count(Token.id)).where(and_(*conditions))
    if join_queue:
        count_query = count_query.join(Queue, Token.queue_id == Queue.id)
    
    total_res = await db.execute(count_query)
    total = total_res.scalar_one()

    # Apply pagination
    query = query.limit(limit).offset(offset)
    result = await db.execute(query)
    
    items = []
    for row in result.all():
        token, q_name, q_prefix, served_first, served_last, served_email, completed_first, completed_last, completed_email = row
        served_name = ""
        if served_first or served_last:
            served_name = f"{served_first or ''} {served_last or ''}".strip()
        elif served_email:
            served_name = served_email.split('@')[0]
            
        completed_name = ""
        if completed_first or completed_last:
            completed_name = f"{completed_first or ''} {completed_last or ''}".strip()
        elif completed_email:
            completed_name = completed_email.split('@')[0]
        
        items.append({
            "id": str(token.id),
            "token_number": token.token_number,
            "queue_name": q_name,
            "queue_prefix": q_prefix,
            "status": token.status.value,
            "customer_name": token.customer_name,
            "customer_phone": token.customer_phone,
            "customer_age": token.customer_age,
            "pax_count": token.pax_count if hasattr(token, 'pax_count') else 1,
            "created_at": token.created_at.isoformat(),
            "served_at": token.served_at.isoformat() if token.served_at else None,
            "completed_at": token.completed_at.isoformat() if token.completed_at else None,
            "called_via_invite": token.called_via_invite,
            "entry_type": getattr(token, "entry_type", "qr"),
            "assigned_line": getattr(token, "assigned_line", None),
            "served_by_staff_name": served_name,
            "completed_by_staff_name": completed_name,
            "removed_by": token.removed_by,
            "deleted_at": token.deleted_at.isoformat() if token.deleted_at else None,
            "skipped_at": token.skipped_at.isoformat() if token.skipped_at else None,
            "recalled_at": token.recalled_at.isoformat() if token.recalled_at else None,
        })

    return {
        "items": items,
        "total": total,
        "limit": limit,
        "offset": offset
    }

async def get_analytics_csv_data(
    db: AsyncSession,
    *,
    org_id: uuid.UUID,
    queue_id: Optional[uuid.UUID] = None,
    session_id: Optional[uuid.UUID] = None,
    search: Optional[str] = None,
    status: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> str:
    """Generate a CSV report of all queue interactions within the date range."""
    import csv
    import io
    from dateutil.parser import parse as parse_date
    from sqlalchemy import and_
    from app.models.queue import Queue
    from app.models.user import User
    from app.models.organization import Organization
    from app.core.tz_helpers import safe_zoneinfo, to_org_local, to_org_local_date, get_org_timezone

    org_tz_str = await get_org_timezone(db, org_id)

    conditions = [Token.org_id == org_id]
    if queue_id:
        conditions.append(Token.queue_id == queue_id)
    if status:
        conditions.append(Token.status == status)
        
    if search:
        from sqlalchemy import or_, cast, String
        search_term = f"%{search.lower()}%"
        conditions.append(or_(
            func.lower(Token.customer_name).like(search_term),
            Token.customer_phone.like(search_term),
            cast(Token.token_number, String).like(search_term)
        ))

    join_queue = False
    if session_id:
        from app.models.session import Session
        sess = await db.scalar(select(Session).where(Session.id == session_id))
        if sess:
            conditions.append(Token.queue_id == sess.queue_id)
            conditions.append(func.date(func.timezone(org_tz_str, Token.created_at)) == sess.session_date)
        else:
            conditions.append(Token.id == None)

    if start_date:
        try:
            from zoneinfo import ZoneInfo
            tz = ZoneInfo(org_tz_str)
            dt = parse_date(start_date)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=tz)
            else:
                dt = dt.astimezone(tz)
            conditions.append(Token.created_at >= dt.astimezone(ZoneInfo("UTC")))
        except Exception:
            pass
            
    if end_date:
        try:
            from zoneinfo import ZoneInfo
            tz = ZoneInfo(org_tz_str)
            ed = parse_date(end_date)
            if ed.tzinfo is None:
                ed = ed.replace(tzinfo=tz)
            else:
                ed = ed.astimezone(tz)
            if ed.hour == 0 and ed.minute == 0 and ed.second == 0:
                ed = ed.replace(hour=23, minute=59, second=59, microsecond=999999)
            conditions.append(Token.created_at <= ed.astimezone(ZoneInfo("UTC")))
        except Exception:
            pass

    from sqlalchemy.orm import aliased
    ServedUser = aliased(User)
    CompletedUser = aliased(User)

    query = select(
        Token,
        Queue.name.label('queue_name'),
        Queue.prefix.label('queue_prefix'),
        ServedUser.first_name.label('served_first'),
        ServedUser.last_name.label('served_last'),
        ServedUser.email.label('served_email'),
        CompletedUser.first_name.label('completed_first'),
        CompletedUser.last_name.label('completed_last'),
        CompletedUser.email.label('completed_email')
    ).outerjoin(Queue, Token.queue_id == Queue.id)\
     .outerjoin(ServedUser, Token.served_by_id == ServedUser.id)\
     .outerjoin(CompletedUser, Token.completed_by_id == CompletedUser.id)\
     .where(and_(*conditions))\
     .order_by(Token.created_at.desc())

    result = await db.execute(query)
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Date", "Token Number", "Queue", "Service Line", "Customer Name", "Customer Phone", "Age", "Pax",
        "Status", "Created At", "Served At", "Completed At", "Skipped At", "Recalled At", "Removed At",
        "Wait Time (mins)", "Serve Time (mins)", "Served By", "Completed By", "Removed By", "Call Method", "Entry Type", "Timezone"
    ])

    for row in result.all():
        token, q_name, q_prefix, served_first, served_last, served_email, completed_first, completed_last, completed_email = row
        
        wait_time_mins = ""
        if token.served_at and token.created_at:
            wait_time_mins = round((token.served_at - token.created_at).total_seconds() / 60.0, 1)
            
        serve_time_mins = ""
        if token.completed_at and token.served_at:
            serve_time_mins = round((token.completed_at - token.served_at).total_seconds() / 60.0, 1)

        served_by = ""
        if served_first or served_last:
            served_by = f"{served_first or ''} {served_last or ''}".strip()
        elif served_email:
            served_by = served_email.split('@')[0]
            
        completed_by = ""
        if completed_first or completed_last:
            completed_by = f"{completed_first or ''} {completed_last or ''}".strip()
        elif completed_email:
            completed_by = completed_email.split('@')[0]

        pax_count_str = str(token.pax_count) if hasattr(token, 'pax_count') else "1"
        service_line = str(getattr(token, 'assigned_line', "")) if getattr(token, 'assigned_line', None) is not None else ""

        entry_type = getattr(token, "entry_type", "qr")
        entry_method = "Manual Entry" if entry_type == "manual" else ("Auto Assigned" if entry_type == "auto" else "QR Code")

        removed_by_label = ""
        if token.removed_by == "customer":
            removed_by_label = "Customer"
        elif token.removed_by == "session_end":
            removed_by_label = "System (Session End)"
        elif token.removed_by:
            removed_by_label = "Staff"

        token_display = f"{q_prefix or ''}{token.token_number}"
        writer.writerow([
            to_org_local_date(token.created_at, org_tz_str),
            token_display,
            q_name or "Unknown",
            service_line,
            token.customer_name or "Walk-in",
            token.customer_phone or "",
            token.customer_age if token.customer_age else "",
            pax_count_str,
            token.status.value if hasattr(token.status, 'value') else str(token.status),
            to_org_local(token.created_at, org_tz_str),
            to_org_local(token.served_at, org_tz_str),
            to_org_local(token.completed_at, org_tz_str),
            to_org_local(token.skipped_at, org_tz_str),
            to_org_local(token.recalled_at, org_tz_str),
            to_org_local(token.deleted_at, org_tz_str),
            wait_time_mins,
            serve_time_mins,
            served_by,
            completed_by,
            removed_by_label,
            "Invite by Number" if token.called_via_invite else "Call Next",
            entry_method,
            org_tz_str
        ])

    return output.getvalue()

import uuid
from typing import Optional, List, Dict, Any
from datetime import datetime, date, timedelta, timezone

from sqlalchemy import select, func, and_, case, extract, text, String, Float
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.organization import Organization
from app.models.token import Token, TokenStatus
from app.models.queue import Queue
from app.models.session import Session
from app.models.user import User

async def get_cross_branch_analytics(
    db: AsyncSession,
    parent_org_id: uuid.UUID,
    branch_id: Optional[uuid.UUID] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
) -> dict:
    from dateutil.parser import parse as parse_date
    from zoneinfo import ZoneInfo
    
    from app.models.parent_organization import ParentOrganization
    parent_org = await db.scalar(select(ParentOrganization).where(ParentOrganization.id == parent_org_id))
    org_tz_str = parent_org.timezone if parent_org and parent_org.timezone else "UTC"
    
    tz = ZoneInfo(org_tz_str)
    
    # 1. Resolve branch filters
    org_query = select(Organization.id, Organization.name, Organization.is_active).where(Organization.parent_organization_id == parent_org_id)
    if branch_id:
        org_query = org_query.where(Organization.id == branch_id)
        
    org_res = await db.execute(org_query)
    org_rows = org_res.all()
    org_ids = [r.id for r in org_rows]
    org_map = {r.id: {"name": r.name, "is_active": r.is_active} for r in org_rows}
    
    if not org_ids:
        # Return empty shell
        return {}

    # 2. Date filters
    token_conditions = [Token.org_id.in_(org_ids)]
    
    if start_date:
        try:
            dt = parse_date(start_date)
            if dt.tzinfo is None: dt = dt.replace(tzinfo=tz)
            else: dt = dt.astimezone(tz)
            token_conditions.append(Token.created_at >= dt.astimezone(ZoneInfo("UTC")))
        except: pass
    if end_date:
        try:
            ed = parse_date(end_date)
            if ed.tzinfo is None: ed = ed.replace(tzinfo=tz)
            else: ed = ed.astimezone(tz)
            if ed.hour == 0 and ed.minute == 0 and ed.second == 0:
                ed = ed.replace(hour=23, minute=59, second=59, microsecond=999999)
            token_conditions.append(Token.created_at <= ed.astimezone(ZoneInfo("UTC")))
        except: pass

    # 3. Base Query for Customer & Time Metrics
    metrics_q = select(
        func.count(Token.id).label("total_customers"),
        func.sum(case((Token.status == TokenStatus.done, 1), else_=0)).label("served"),
        func.sum(case((Token.status == TokenStatus.waiting, 1), else_=0)).label("waiting"),
        func.sum(case((Token.status.in_([TokenStatus.skipped, TokenStatus.deleted]), 1), else_=0)).label("abandoned"),
        func.avg(func.extract('epoch', Token.served_at - Token.created_at)).label('avg_wait_sec'),
        func.avg(func.extract('epoch', Token.completed_at - Token.served_at)).label('avg_serve_sec'),
    ).where(and_(*token_conditions))
    
    m_res = await db.execute(metrics_q)
    m_row = m_res.first()
    
    total_customers = m_row.total_customers or 0
    served = int(m_row.served or 0)
    waiting = int(m_row.waiting or 0)
    abandoned = int(m_row.abandoned or 0)
    
    completion_rate = f"{round((served / total_customers * 100) if total_customers > 0 else 0, 1)}%"
    abandonment_rate = f"{round((abandoned / total_customers * 100) if total_customers > 0 else 0, 1)}%"
    
    def format_time(seconds: float | None) -> str:
        if not seconds: return "00:00:00"
        m, s = divmod(int(seconds), 60)
        h, m = divmod(m, 60)
        return f"{h:02d}:{m:02d}:{s:02d}"
        
    avg_wait = format_time(m_row.avg_wait_sec)
    avg_serve = format_time(m_row.avg_serve_sec)
    
    # 4. Operations Metrics (we query Queues and Sessions that belong to org_ids)
    q_res = await db.execute(select(func.count(Queue.id)).where(Queue.org_id.in_(org_ids), Queue.is_active == True))
    active_queues = q_res.scalar_one() or 0
    
    # Actually, sessions are active if they are today, but let's just count active queues and total active branches.
    active_branches = sum(1 for o in org_map.values() if o["is_active"])
    
    today = datetime.now(timezone.utc).date()
    s_res = await db.execute(select(func.count(Session.id)).where(Session.org_id.in_(org_ids), Session.session_date == today))
    active_sessions = s_res.scalar_one() or 0
    
    # Calculate operated queues (distinct queues that handled tokens during this period)
    oq_res = await db.execute(select(func.count(func.distinct(Token.queue_id))).where(and_(*token_conditions)))
    operated_queues = oq_res.scalar_one() or 0
    
    # Calculate online staff (active in the last hour)
    from datetime import timedelta
    from app.models.user import User
    one_hour_ago = datetime.now(timezone.utc) - timedelta(hours=1)
    os_res = await db.execute(select(func.count(User.id)).where(User.org_id.in_(org_ids), User.last_active_at >= one_hour_ago))
    online_staff = os_res.scalar_one() or 0
    
    # 5. Volume Trend (Daily)
    trend_q = select(
        func.date(func.timezone(org_tz_str, Token.created_at)).label('dt'),
        func.count(Token.id).label("served")
    ).where(and_(*token_conditions, Token.status == TokenStatus.done)).group_by('dt').order_by('dt')
    
    trend_res = await db.execute(trend_q)
    volume_trend = [{"date": r.dt.isoformat(), "customers_served": r.served} for r in trend_res.all()]
    
    # 6. Branch Ranking
    branch_q = select(
        Token.org_id,
        func.count(Token.id).label("total"),
        func.sum(case((Token.status == TokenStatus.done, 1), else_=0)).label("served"),
        func.avg(func.extract('epoch', Token.served_at - Token.created_at)).label('avg_wait_sec'),
        func.avg(func.extract('epoch', Token.completed_at - Token.served_at)).label('avg_serve_sec'),
    ).where(and_(*token_conditions)).group_by(Token.org_id)
    
    branch_res = await db.execute(branch_q)
    
    branch_ranking_unsorted = []
    for r in branch_res.all():
        b_total = r.total or 0
        b_served = int(r.served or 0)
        cr = (b_served / b_total * 100) if b_total > 0 else 0
        
        # Health score: 60% completion rate, 40% wait time (inverse, assuming < 30m is good)
        wt = float(r.avg_wait_sec or 0)
        wt_score = max(0, 100 - (wt / 1800 * 100)) # 30 mins = 0 score
        health = int((cr * 0.6) + (wt_score * 0.4))
        
        h_status = "Excellent" if health >= 95 else "Good" if health >= 80 else "Warning" if health >= 60 else "Critical"
        
        branch_ranking_unsorted.append({
            "branch": org_map[r.org_id]["name"],
            "customers_served": b_served,
            "avg_wait_time": format_time(r.avg_wait_sec),
            "avg_service_time": format_time(r.avg_serve_sec),
            "raw_wait_sec": float(r.avg_wait_sec or 0),
            "completion_rate": f"{round(cr, 1)}%",
            "health_score": health,
            "health_status": h_status
        })
        
    branch_ranking_unsorted.sort(key=lambda x: (x["health_score"], x["customers_served"]), reverse=True)
    branch_ranking = []
    for i, b in enumerate(branch_ranking_unsorted):
        b["rank"] = i + 1
        branch_ranking.append(b)
        
    # 7. Queue Analytics
    queue_q = select(
        Queue.name,
        Token.org_id,
        func.count(Token.id).label("served"),
        func.avg(func.extract('epoch', Token.served_at - Token.created_at)).label('avg_wait_sec'),
        func.avg(func.extract('epoch', Token.completed_at - Token.served_at)).label('avg_serve_sec'),
    ).join(Queue, Token.queue_id == Queue.id).where(and_(*token_conditions, Token.status == TokenStatus.done)).group_by(Queue.id, Token.org_id).order_by(func.count(Token.id).desc())
    
    queue_res = await db.execute(queue_q)
    queue_analytics = []
    for r in queue_res.all():
        queue_analytics.append({
            "queue_name": r.name,
            "branch": org_map[r.org_id]["name"],
            "customers_served": r.served,
            "avg_wait_time": format_time(r.avg_wait_sec),
            "avg_service_time": format_time(r.avg_serve_sec)
        })
        
    # 8. Peak Traffic (Hourly distribution)
    peak_q = select(
        func.extract('hour', func.timezone(org_tz_str, Token.created_at)).label('hr'),
        func.count(Token.id).label("arrived"),
        func.sum(case((Token.status == TokenStatus.done, 1), else_=0)).label("served"),
        func.avg(func.extract('epoch', Token.served_at - Token.created_at)).label('avg_wait_sec'),
    ).where(and_(*token_conditions)).group_by('hr').order_by('hr')
    
    peak_res = await db.execute(peak_q)
    peak_traffic = []
    max_arrived = 0
    peak_hr = None
    
    rows = peak_res.all()
    for r in rows:
        if r.arrived > max_arrived:
            max_arrived = r.arrived
            peak_hr = r.hr
            
    # Format peak hour for TimeMetrics
    formatted_peak = "-"
    if peak_hr is not None:
        hr_int = int(peak_hr)
        ampm = "AM" if hr_int < 12 else "PM"
        display_hr = hr_int if hr_int <= 12 else hr_int - 12
        if display_hr == 0: display_hr = 12
        formatted_peak = f"{display_hr}:00 {ampm} - {display_hr}:59 {ampm}"
        
    for r in rows:
        hr_int = int(r.hr)
        ampm = "AM" if hr_int < 12 else "PM"
        display_hr = hr_int if hr_int <= 12 else hr_int - 12
        if display_hr == 0: display_hr = 12
        # Convert avg wait seconds to minutes (rounded to 1dp)
        avg_wait_min = round(float(r.avg_wait_sec) / 60, 1) if r.avg_wait_sec else 0
        
        peak_traffic.append({
            "time_block": f"{display_hr}:00 {ampm}",
            "customers_arrived": r.arrived,
            "customers_served": int(r.served or 0),
            "avg_wait_minutes": avg_wait_min,
            "is_peak": (r.hr == peak_hr)
        })
        
    # 9. Staff Performance
    staff_q = select(
        User.first_name, User.last_name, User.email,
        Token.org_id,
        func.count(Token.id).label("served"),
        func.avg(func.extract('epoch', Token.completed_at - Token.served_at)).label('avg_serve_sec'),
    ).join(User, Token.served_by_id == User.id).where(and_(*token_conditions, Token.status == TokenStatus.done)).group_by(User.id, Token.org_id).order_by(func.count(Token.id).desc())
    
    staff_res = await db.execute(staff_q)
    staff_performance = []
    for r in staff_res.all():
        name = f"{r.first_name or ''} {r.last_name or ''}".strip()
        if not name: name = r.email.split('@')[0]
        staff_performance.append({
            "staff_name": name,
            "branch": org_map[r.org_id]["name"],
            "customers_served": r.served,
            "avg_service_time": format_time(r.avg_serve_sec)
        })
        
    # 10. Generate Insights
    insights = []
    
    abandonment_pct = (abandoned / total_customers * 100) if total_customers > 0 else 0
    if abandonment_pct > 10:
        insights.append(f"🚨 Insight: You experienced a {round(abandonment_pct, 1)}% abandonment rate today. This correlates heavily with peak traffic bottlenecks.")
        
    network_avg_wait = float(m_row.avg_wait_sec or 0)
    if network_avg_wait > 0 and branch_ranking:
        for b in branch_ranking:
            if b['raw_wait_sec'] > (network_avg_wait * 1.5) and b['raw_wait_sec'] > 300: # at least 5 mins and 1.5x avg
                insights.append(f"🚨 Warning: {b['branch']} has an average wait time of {b['avg_wait_time']}, which is significantly higher than the network average.")
                break # Just alert the worst offender
                
    if not insights:
        if branch_ranking:
            top_branch = branch_ranking[0]
            if top_branch['customers_served'] > 0:
                insights.append(f"💡 All systems normal. {top_branch['branch']} is leading with {top_branch['customers_served']} customers served and a {top_branch['completion_rate']} completion rate.")
            else:
                insights.append(f"💡 {top_branch['branch']} is currently the most active branch, though service completion data is still accumulating.")
        else:
            insights.append("💡 AI is currently collecting data. Strategic insights will be generated automatically as queue activity begins.")

    # 11. PAX / Group Size Analytics
    pax_q = select(
        func.cast(Token.pax_count, String).label('group_size'),
        func.count(Token.id).label('token_count'),
        func.sum(Token.pax_count).label('total_pax'),
        func.avg(func.extract('epoch', Token.served_at - Token.created_at)).label('avg_wait_sec'),
        func.avg(func.extract('epoch', Token.completed_at - Token.served_at)).label('avg_serve_sec'),
    ).where(and_(*token_conditions)).group_by(Token.pax_count).order_by(Token.pax_count)

    pax_res = await db.execute(pax_q)
    pax_analytics = []
    for r in pax_res.all():
        pax_analytics.append({
            "group_size": str(r.group_size),
            "token_count": r.token_count,
            "total_pax": int(r.total_pax or 0),
            "avg_wait_time": format_time(r.avg_wait_sec),
            "avg_service_time": format_time(r.avg_serve_sec),
        })

    # PAX summary
    pax_summary_q = select(
        func.sum(Token.pax_count).label('total_headcount'),
        func.avg(func.cast(Token.pax_count, Float)).label('avg_group_size'),
        func.max(Token.pax_count).label('largest_group'),
    ).where(and_(*token_conditions))
    pax_summary_res = await db.execute(pax_summary_q)
    pax_summary_row = pax_summary_res.first()
    pax_summary = {
        "total_headcount": int(pax_summary_row.total_headcount or 0),
        "avg_group_size": round(float(pax_summary_row.avg_group_size or 1.0), 1),
        "largest_group": int(pax_summary_row.largest_group or 1),
    } if pax_summary_row and pax_summary_row.total_headcount else None

    return {
        "customer_metrics": {
            "total_customers": total_customers,
            "customers_served": served,
            "customers_waiting": waiting,
            "customers_abandoned": abandoned,
            "completion_rate": completion_rate,
            "abandonment_rate": abandonment_rate
        },
        "time_metrics": {
            "avg_wait_time": avg_wait,
            "avg_service_time": avg_serve,
            "peak_hour": formatted_peak
        },
        "operations_metrics": {
            "active_branches": active_branches,
            "active_sessions": active_sessions,
            "active_queues": active_queues,
            "operated_queues": operated_queues,
            "online_staff": online_staff
        },
        "volume_trend": volume_trend,
        "branch_ranking": branch_ranking,
        "queue_analytics": queue_analytics,
        "peak_traffic": peak_traffic,
        "staff_performance": staff_performance,
        "pax_analytics": pax_analytics,
        "pax_summary": pax_summary,
        "insights": insights
    }

import uuid
from typing import Optional, List
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func

from app.models.token import Token, TokenStatus
from app.models.queue import Queue
from app.models.user import User
from app.models.organization import Organization

async def get_cross_branch_excel_data(
    db: AsyncSession,
    org_ids: List[uuid.UUID],
    parent_org_id: Optional[uuid.UUID] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> bytes:
    import csv
    import io
    from dateutil.parser import parse as parse_date
    from zoneinfo import ZoneInfo
    from app.models.organization import Organization
    
    # We use UTC for multi-org CSV exports by default, or we could fetch the first org's tz.
    # For simplicity, we just use UTC here or default to 'Asia/Kolkata' if you really want.
    org_tz_str = "UTC"
    if org_ids:
        first_org = await db.scalar(select(Organization).where(Organization.id == org_ids[0]))
        if first_org and first_org.timezone:
            org_tz_str = first_org.timezone

    tz = ZoneInfo(org_tz_str)
    
    conditions = [Token.org_id.in_(org_ids)]
    
    if start_date:
        try:
            dt = parse_date(start_date)
            if dt.tzinfo is None: dt = dt.replace(tzinfo=tz)
            else: dt = dt.astimezone(tz)
            conditions.append(Token.created_at >= dt.astimezone(ZoneInfo("UTC")))
        except: pass
        
    if end_date:
        try:
            ed = parse_date(end_date)
            if ed.tzinfo is None: ed = ed.replace(tzinfo=tz)
            else: ed = ed.astimezone(tz)
            if ed.hour == 0 and ed.minute == 0 and ed.second == 0:
                ed = ed.replace(hour=23, minute=59, second=59, microsecond=999999)
            conditions.append(Token.created_at <= ed.astimezone(ZoneInfo("UTC")))
        except: pass

    from sqlalchemy.orm import aliased
    ServedUser = aliased(User)
    CompletedUser = aliased(User)

    query = select(
        Token,
        Queue.name.label('queue_name'),
        Organization.name.label('branch_name'),
        ServedUser.first_name.label('served_first'),
        ServedUser.last_name.label('served_last'),
        ServedUser.email.label('served_email'),
        CompletedUser.first_name.label('completed_first'),
        CompletedUser.last_name.label('completed_last'),
        CompletedUser.email.label('completed_email')
    ).outerjoin(Queue, Token.queue_id == Queue.id)\
     .outerjoin(Organization, Token.org_id == Organization.id)\
     .outerjoin(ServedUser, Token.served_by_id == ServedUser.id)\
     .outerjoin(CompletedUser, Token.completed_by_id == CompletedUser.id)\
     .where(and_(*conditions))\
     .order_by(Token.created_at.desc())

    # Build the Dashboard Summary first
    from app.services.analytics_service import get_cross_branch_analytics
    
    analytics_summary = None
    if parent_org_id:
        # Since get_cross_branch_excel_data now takes parent_org_id, we can fetch summary
        # Note: If branch filtering was active, we should pass branch_id instead if we only had one.
        # But we only have org_ids. Let's just use parent_org_id and let the analytics function filter if needed.
        # Actually, if org_ids is length 1, that's the branch.
        b_id = org_ids[0] if len(org_ids) == 1 else None
        
        analytics_summary = await get_cross_branch_analytics(
            db=db,
            parent_org_id=parent_org_id,
            branch_id=b_id,
            start_date=start_date,
            end_date=end_date
        )

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    
    wb = Workbook()
    
    # --- Tab 1: Dashboard Summary ---
    ws_summary = wb.active
    ws_summary.title = "Dashboard Summary"
    
    ws_summary.append(["Q4Queue Executive Dashboard Summary"])
    ws_summary.append([])
    
    ws_summary["A1"].font = Font(bold=True, size=16)
    
    if analytics_summary:
        metrics = analytics_summary.get("customer_metrics", {})
        time_metrics = analytics_summary.get("time_metrics", {})
        
        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["Total Served", metrics.get("customers_served", 0)])
        ws_summary.append(["Total Abandoned", metrics.get("customers_abandoned", 0)])
        ws_summary.append(["Completion Rate", metrics.get("completion_rate", "0%")])
        ws_summary.append(["Abandonment Rate", metrics.get("abandonment_rate", "0%")])
        ws_summary.append(["Average Wait Time", time_metrics.get("avg_wait_time", "00:00:00")])
        ws_summary.append(["Average Service Time", time_metrics.get("avg_service_time", "00:00:00")])
        ws_summary.append(["Peak Hour", time_metrics.get("peak_hour", "-")])
    else:
        ws_summary.append(["Summary Data", "Not Available"])

    # Make bold
    for row in ws_summary.iter_rows(min_row=3, max_row=10, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)
            
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20

    # --- Tab 2: Raw Data ---
    ws_raw = wb.create_sheet(title="Raw Data")
    
    headers = [
        "Date", "Branch", "Token Number", "Queue", "Customer Name", "Customer Phone", "Age", "Pax",
        "Status", "Created At", "Served At", "Completed At", "Skipped At", "Recalled At", "Removed At",
        "Wait Time (mins)", "Serve Time (mins)", "Served By", "Completed By", "Removed By", "Call Method", "Entry Type"
    ]
    ws_raw.append(headers)
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, len(headers) + 1):
        cell = ws_raw.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        ws_raw.column_dimensions[get_column_letter(col_idx)].width = 22

    result = await db.stream(query)
    
    async for row in result:
        token, q_name, b_name, served_first, served_last, served_email, completed_first, completed_last, completed_email = row
        
        wait_time_mins = ""
        if token.served_at and token.created_at:
            wait_time_mins = round((token.served_at - token.created_at).total_seconds() / 60.0, 1)
            
        serve_time_mins = ""
        if token.completed_at and token.served_at:
            serve_time_mins = round((token.completed_at - token.served_at).total_seconds() / 60.0, 1)

        served_by = ""
        if served_first or served_last:
            served_by = f"{served_first or ''} {served_last or ''}".strip()
        elif served_email:
            served_by = served_email.split('@')[0]

        completed_by = ""
        if completed_first or completed_last:
            completed_by = f"{completed_first or ''} {completed_last or ''}".strip()
        elif completed_email:
            completed_by = completed_email.split('@')[0]

        removed_by_label = ""
        if token.removed_by == "customer":
            removed_by_label = "Customer"
        elif token.removed_by == "session_end":
            removed_by_label = "System (Session End)"
        elif token.removed_by:
            removed_by_label = "Staff"

        # Convert timestamps to local org timezone
        created_local = token.created_at.astimezone(tz) if token.created_at else None
        served_local = token.served_at.astimezone(tz) if token.served_at else None
        completed_local = token.completed_at.astimezone(tz) if token.completed_at else None
        skipped_local = token.skipped_at.astimezone(tz) if token.skipped_at else None
        recalled_local = token.recalled_at.astimezone(tz) if token.recalled_at else None
        deleted_local = token.deleted_at.astimezone(tz) if token.deleted_at else None

        ws_raw.append([
            created_local.strftime("%Y-%m-%d") if created_local else "",
            b_name or "Unknown",
            token.token_number,
            q_name or "Unknown",
            token.customer_name or "Walk-in",
            token.customer_phone or "",
            token.customer_age if token.customer_age else "",
            str(token.pax_count) if hasattr(token, 'pax_count') else "1",
            token.status.value,
            created_local.isoformat() if created_local else "",
            served_local.isoformat() if served_local else "",
            completed_local.isoformat() if completed_local else "",
            skipped_local.isoformat() if skipped_local else "",
            recalled_local.isoformat() if recalled_local else "",
            deleted_local.isoformat() if deleted_local else "",
            wait_time_mins,
            serve_time_mins,
            served_by,
            completed_by,
            removed_by_label,
            "Invite by Number" if token.called_via_invite else "Call Next",
            getattr(token, "entry_type", "qr").title()
        ])
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
